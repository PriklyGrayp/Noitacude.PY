import os
from openpyxl import Workbook, load_workbook


def parse_xlsx(in_file, out_file, out_start_row, mirror_table, out_filters=False):
    '''
    This function opens input xlsx files and places them i nthe output file with

    Args:
        in_file: xlsx file. The File to source data from.
        out_file: xlsx file. The file to apply the sorced data too.
        out_start_row: Int. The number of the headder row from the out_file.
        mirror_table: Dict. The mirror table from the in_file to the out_file.
        out_filters: Dict. Filters for the sheet. Default is False.
    '''
    # Open the xlsx input file
    in_workbook = load_workbook(filename=in_file)
    in_sheet = in_workbook.active

    # Create the xlsx output file
    if os.path.exists(out_file):
        out_workbook = load_workbook(filename=out_file)
    else:
        out_workbook = Workbook()
    out_sheet = out_workbook.active

    # Parse and reorder
    for in_column in in_sheet.iter_cols(values_only=True):

        # Validate the column
        in_title = in_column[0]
        if not in_title in mirror_table:
            continue
        out_title = mirror_table[in_title]

        # Set out cell values
        out_row_no = out_start_row
        for in_value in in_column:
            out_sheet["{}{}".format(out_title, out_row_no)] = in_value
            out_row_no += 1

    # Apply filters
    if out_filters:
        apply_filters(out_sheet, out_filters)

    # Save the output xlsx file
    out_workbook.save(filename=out_file)


def apply_filters(out_sheet, out_filters):
    '''
    Applies filters to the sheet

    Args:
        out_sheet: Obj. The sheet to apply the filters to.
        out_filters: Dict. Filters for the sheet.
    '''
    for out_cell, out_filter in out_filters.items():
        out_sheet.auto_filter.ref("{}:{}".format(out_cell, out_filter))


out_start_row = 1
out_filters = False
mirror_table = {1:"A", 2:"B", 3:"C", 4:"D", 5:"E", 6:"F"}  # Input : Output
out_file = "C:/Users/mvaar/Documents/GitHub/Noitacude.PY/Excel_Parse/test/Output.xlsx"
in_files = ["C:/Users/mvaar/Documents/GitHub/Noitacude.PY/Excel_Parse/test/InputA.xlsx", "C:/Users/mvaar/Documents/GitHub/Noitacude.PY/Excel_Parse/test/InputA.xlsx"]


for in_file in in_files:
    parse_xlsx(in_file, out_file, out_start_row, mirror_table, out_filters)
